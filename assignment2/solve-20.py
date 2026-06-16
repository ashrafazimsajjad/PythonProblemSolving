from openpyxl import load_workbook

# Load existing Excel file
wb = load_workbook("data.xlsx")
ws = wb.active

# Read a specific cell (B2 = Age of Sajjad)
print("Original Value (B2):", ws["B2"].value)
print("Original Value (C3):", ws["C3"].value)

# Update the cell value
ws["B2"] = 99
ws["C3"] = "Barishal"

# Save the updated file
wb.save("updated_data.xlsx")

print("Cell updated successfully!")